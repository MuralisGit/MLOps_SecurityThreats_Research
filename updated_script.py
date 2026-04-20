import time
import re
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook, load_workbook

# -----------------------------
# CONFIGURATION
# -----------------------------
GITHUB_TOKEN = "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"
INPUT_FILE = "filtered_full_6_stages.xlsx"
REPO_COLUMN = "Repository"

OUTPUT_FILE = "issues_extracted_last_3_years_all_repos.xlsx"
SUMMARY_FILE = "repo_processing_summary.xlsx"

PER_PAGE = 100
EXCEL_CHAR_LIMIT = 32767

HEADERS = {
    "Authorization": f"Bearer {GITHUB_TOKEN}",
    "Accept": "application/vnd.github+json"
}

OUTPUT_COLUMNS = [
    "repo",
    "issue_number",
    "title",
    "state",
    "created_at",
    "closed_at",
    "author",
    "labels",
    "comments_count",
    "issue_body",
    "comments_text",
    "url",
]

SUMMARY_COLUMNS = [
    "repo",
    "issues_count",
    "status",
    "processed_at_utc",
    "error_message",
]

_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]")

# -----------------------------
# HELPER: HANDLE RATE LIMIT / RETRIES
# -----------------------------
def safe_get(url, params=None, max_retries=10):
    retries = 0

    while retries < max_retries:
        try:
            response = requests.get(url, headers=HEADERS, params=params, timeout=60)

            if response.status_code == 200:
                return response

            if response.status_code == 403:
                reset_time = response.headers.get("X-RateLimit-Reset")
                if reset_time:
                    wait_seconds = max(1, int(reset_time) - int(time.time()) + 5)
                else:
                    wait_seconds = 60
                print(f"⏳ Rate limit reached. Sleeping {wait_seconds} sec...")
                time.sleep(wait_seconds)
                continue

            if response.status_code in (500, 502, 503, 504):
                retries += 1
                sleep_for = 5 * retries
                print(f"⚠️ Server error {response.status_code}. Retrying in {sleep_for} sec...")
                time.sleep(sleep_for)
                continue

            print(f"❌ Error {response.status_code}: {response.text}")
            return None

        except requests.exceptions.RequestException as e:
            retries += 1
            sleep_for = 5 * retries
            print(f"⚠️ Request failed: {e}. Retrying in {sleep_for} sec...")
            time.sleep(sleep_for)

    print("❌ Max retries exceeded")
    return None

# -----------------------------
# FETCH COMMENTS
# -----------------------------
def fetch_comments(comments_url):
    response = safe_get(comments_url)

    if not response:
        return []

    comments_data = response.json()

    if not isinstance(comments_data, list):
        return []

    return [
        c.get("body", "")
        for c in comments_data
        if isinstance(c, dict)
    ]

# -----------------------------
# DATE RANGE SPLIT
# -----------------------------
def generate_date_ranges():
    today = datetime.now()
    three_years_ago = today - timedelta(days=365 * 3)

    ranges = []
    current_start = three_years_ago

    while current_start <= today:
        current_end = current_start + timedelta(days=89)
        if current_end > today:
            current_end = today

        ranges.append((
            current_start.strftime("%Y-%m-%d"),
            current_end.strftime("%Y-%m-%d")
        ))

        current_start = current_end + timedelta(days=1)

    return ranges

# -----------------------------
# FETCH ISSUES FOR ONE REPO
# returns (defects, error_message)
# -----------------------------
def fetch_all_defects(repo):
    defects = []
    seen_issue_numbers = set()
    date_ranges = generate_date_ranges()
    stop_all = False # flag to stop all processing if critical error occurs

    for start_date, end_date in date_ranges:
        print(f"\n📅 [{repo}] Fetching issues created from {start_date} to {end_date}")

        if stop_all:
            break

        page = 1

        while True:
            url = "https://api.github.com/search/issues"
            params = {
                "q": f"repo:{repo} type:issue created:{start_date}..{end_date}",
                "per_page": PER_PAGE,
                "page": page
            }

            response = safe_get(url, params)

            if not response:
                error_message = f"safe_get returned None for range {start_date}..{end_date}, page {page}"
                print(f"❌ safe_get failed — stopping all loops for repo: [{repo}] {error_message}")
                stop_all = True
                return defects, error_message

            data = response.json()
            issues = data.get("items", [])

            if not issues:
                break

            for issue in issues:
                issue_number = issue["number"]
                unique_key = f"{repo}#{issue_number}"

                if unique_key in seen_issue_numbers:
                    continue
                seen_issue_numbers.add(unique_key)

                comments = []
                if issue.get("comments", 0) > 0:
                    comments = fetch_comments(issue["comments_url"])

                defects.append({
                    "repo": repo,
                    "issue_number": issue_number,
                    "title": issue.get("title", ""),
                    "state": issue.get("state", ""),
                    "created_at": issue.get("created_at"),
                    "closed_at": issue.get("closed_at"),
                    "author": (issue.get("user") or {}).get("login", ""),
                    "labels": ", ".join(
                        label.get("name", "")
                        for label in issue.get("labels", [])
                    ),
                    "comments_count": issue.get("comments", 0),
                    "issue_body": issue.get("body", ""),
                    "comments_text": " || ".join(comments),
                    "url": issue.get("html_url", "")
                })

            print(f"➡️ [{repo}] Page {page} | collected {len(defects)} issues so far")

            if len(issues) < PER_PAGE:
                break

            page += 1
            time.sleep(1)

    return defects, None

# -----------------------------
# CLEAN EXCEL TEXT
# -----------------------------
def clean_excel_text(value):
    if pd.isna(value):
        return value

    if not isinstance(value, str):
        value = str(value)

    value = _ILLEGAL_XLSX_RE.sub("", value)

    if len(value) > EXCEL_CHAR_LIMIT:
        value = value[:EXCEL_CHAR_LIMIT]

    return value

# -----------------------------
# EXCEL INIT / APPEND
# -----------------------------
def initialize_workbook(filename, sheet_name, columns):
    path = Path(filename)

    if path.exists():
        print(f"📄 File already exists: {filename}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)
    wb.save(filename)
    wb.close()
    print(f"📄 Created workbook: {filename}")

def append_rows_to_excel(data, filename, sheet_name, columns):
    if not data:
        print(f"⚠️ No rows to append to {filename}")
        return

    df = pd.DataFrame(data)

    for col in columns:
        if col not in df.columns:
            df[col] = ""

    df = df[columns]

    for col in df.columns:
        df[col] = df[col].map(clean_excel_text)

    wb = load_workbook(filename)
    ws = wb[sheet_name]

    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

    wb.save(filename)
    wb.close()

    print(f"📁 Appended {len(df)} rows to {filename}")

# -----------------------------
# LOAD REPOS FROM INPUT FILE
# -----------------------------
def load_repos_from_excel(input_file, repo_column):
    df = pd.read_excel(input_file)

    if repo_column not in df.columns:
        raise ValueError(
            f"Column '{repo_column}' not found. Available columns: {list(df.columns)}"
        )

    repos = (
        df[repo_column]
        .dropna()
        .astype(str)
        .str.strip()
    )

    repos = repos[repos.str.contains(r"^[^/]+/[^/]+$", regex=True)]

    return repos.drop_duplicates().tolist()

# -----------------------------
# RUN
# -----------------------------
repos = load_repos_from_excel(INPUT_FILE, REPO_COLUMN)
print(f"🔍 Found {len(repos)} repositories in {INPUT_FILE}")

initialize_workbook(OUTPUT_FILE, "Issues", OUTPUT_COLUMNS)
initialize_workbook(SUMMARY_FILE, "RepoSummary", SUMMARY_COLUMNS)

for idx, repo_name in enumerate(repos, start=1):
    print(f"\n{'=' * 80}")
    print(f"[{idx}/{len(repos)}] Processing {repo_name}")
    print(f"{'=' * 80}")

    processed_at_utc = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

    try:
        repo_defects, error_message = fetch_all_defects(repo_name)

        if repo_defects:
            append_rows_to_excel(
                repo_defects,
                OUTPUT_FILE,
                "Issues",
                OUTPUT_COLUMNS
            )
            print(f"\n✅ Appended issues for {repo_name}")

        summary_row = [{
            "repo": repo_name,
            "issues_count": len(repo_defects),
            "status": "FAILED" if error_message else "SUCCESS",
            "processed_at_utc": processed_at_utc,
            "error_message": error_message or ""
        }]

        append_rows_to_excel(
            summary_row,
            SUMMARY_FILE,
            "RepoSummary",
            SUMMARY_COLUMNS
        )

    except Exception as e:
        summary_row = [{
            "repo": repo_name,
            "issues_count": 0,
            "status": "FAILED",
            "processed_at_utc": processed_at_utc,
            "error_message": str(e)
        }]

        append_rows_to_excel(
            summary_row,
            SUMMARY_FILE,
            "RepoSummary",
            SUMMARY_COLUMNS
        )

        print(f"❌ [{repo_name}] Unexpected error: {e}")

    time.sleep(2)

print("\n✅ Finished processing all repositories")